from __future__ import annotations

from collections import defaultdict
from copy import copy
from datetime import date, datetime
from io import BytesIO
import os
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, Query

from auth.dependencies import require_module
from database import run_query
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


router = APIRouter(dependencies=[Depends(require_module("porsche"))])

# Fuente unica de Porsche: vista consolidada, no tablas de staging.
DASHBOARD_VIEW = "dbo.dashboard_data"
PW_METAS_TABLE = "dbo.tmp_PW_metas"
PW_CIERRE_TRAMOS = ("31-60", "61-90", "90+")

TRAMO_ORDER = ["31-60", "61-90", "91-120", "121-150", "151-180", "181-210", "211-240"]
EXPORT_TRAMO_ORDER = ["31-60", "61-90", "91-120", "121-150", "151-180", "181-210", "211-240", "241-270"]
CONTACTO_META = {"31-60": 0.80, "61-90": 0.70, "91-120": 0.50, "121-150": 0.50, "151-180": 0.50, "181-210": 0.50, "211-240": 0.50}
RECUPERACION_META = {"31-60": 0.70, "61-90": 0.60, "91-120": 0.50, "121-150": 0.50, "151-180": 0.50, "181-210": 0.50, "211-240": 0.50}
PORSCHE_EXPORT_TEMPLATE_ENV = "PORSCHE_EXPORT_TEMPLATE"

EXPORT_DASHBOARD_CONFIG = {
    "contactabilidad": {"start_row": 11, "total_row": 19, "assigned_key": "asignado", "value_key": "casos_contactados", "pct_key": "pct_contacto"},
    "promesas_pago": {"start_row": 24, "total_row": 32, "assigned_key": "asignado", "value_key": "casos_con_promesa", "pct_key": "pct_promesa_pago"},
    "promesas_cumplidas": {"start_row": 37, "total_row": 45, "assigned_key": "asignado", "value_key": "promesas_cumplidas", "pct_key": "pct_cumplimiento_promesa"},
    "promesas_incumplidas": {"start_row": 50, "total_row": 58, "assigned_key": "asignado", "value_key": "promesas_incumplidas", "pct_key": "pct_incumplido"},
    "recuperacion": {"start_row": 63, "total_row": 71, "assigned_key": "asignado", "value_key": "casos_pagados", "pct_key": "pct_recupero"},
    "contenido": {"start_row": 76, "total_row": 84, "assigned_key": "asignado", "value_key": "casos_contenidos", "pct_key": "pct_contenido"},
    "normalizado": {"start_row": 89, "total_row": 97, "assigned_key": "asignado", "value_key": "casos_normalizados", "pct_key": "pct_normalizado"},
    "campana_renegociacion": {"start_row": 102, "total_row": 104, "labels": ["Solicitudes", "Efectivas"], "assigned_key": "asignado", "value_key": "kpi", "pct_key": "pct_kpi"},
    "tpr": {"start_row": 109, "total_row": 117, "assigned_key": "asignado", "value_key": "TPR", "pct_key": "pct_kpi"},
    "reiteracion_contacto": {"start_row": 122, "total_row": 130, "assigned_key": "casos_sin_contacto", "value_key": "RC", "pct_key": "pct_kpi"},
}

DATA_EXPORT_COLUMNS = [
    ("op", "numero_operacion", "nro_operacion", "contrato"),
    ("rut", "rut_cliente"),
    ("rut_deudor", "rut_numero", "rut"),
    ("tramo", "negocio_pw"),
    ("deuda", "monto_adeudado", "total_deuda"),
    ("accion", "contacto_gestion", "gestion_accion"),
    ("gestion", "respuesta_gestion", "sub_gestion", "contactogestion"),
    ("respuesta_gestion", "observaciones", "respuesta", "obs", "respuestagestion"),
    ("ult_fecha_gestion", "ultima_fecha_gestion", "fecha_gestion", "gestionfecha"),
    ("hora_gestion", "gestion_hora", "gestionhora"),
    ("marca_rene",),
    ("pagos", "total_pagado", "monto_pagado"),
    ("contenido",),
    ("normaliza", "normalizado"),
    ("fecha_gestion_compromiso", "fecha_compromiso"),
    ("filtro_compromiso",),
    ("filtro_pago", "pago_bin"),
    ("filtro_no_pago", "incumplido"),
    ("pago_con_compromiso",),
    ("rene_ofrecida",),
    ("q_contacto",),
    ("contactabilidad",),
    ("rene_cursada",),
    ("intensidad",),
    ("dias_habiles", "dias"),
]

EXPORT_VISIBLE_SHEETS = ("Dashboard", "DATA")


def _coerce_datetime(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)

    text = str(value).strip()
    if not text:
        return None

    formats = (
        "%Y-%m-%d %H:%M:%S.%f",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%m-%Y",
        "%m/%Y",
    )
    for fmt in formats:
        try:
            parsed = datetime.strptime(text, fmt)
            if fmt in ("%m-%Y", "%m/%Y"):
                return datetime(parsed.year, parsed.month, 1)
            return parsed
        except ValueError:
            continue

    try:
        return datetime.fromisoformat(text.replace("Z", ""))
    except ValueError:
        return None


def _month_label(value) -> str:
    parsed = _coerce_datetime(value)
    if parsed is None:
        return ""
    return f"{parsed.month:02d}-{parsed.year:04d}"


def _month_sort_key(label: str) -> tuple[int, int]:
    try:
        month_str, year_str = label.split("-", 1)
        return int(year_str), int(month_str)
    except Exception:
        return 0, 0


def _selected_month_rows(rows: list[dict], mes: str | None) -> list[dict]:
    if not mes:
        return rows

    return [row for row in rows if _month_label(row.get("fecha_hora_carga")) == mes]


def _period_start_from_month_label(label: str | None) -> date | None:
    text = str(label or "").strip()
    if not text:
        return None

    for fmt in ("%m-%Y", "%Y-%m", "%Y-%m-%d", "%m/%Y"):
        try:
            parsed = datetime.strptime(text, fmt)
            return date(parsed.year, parsed.month, 1)
        except ValueError:
            continue

    return None


def _normalize_pw_tramo(value: str | None) -> str:
    text = str(value or "").strip().upper()
    if text in PW_CIERRE_TRAMOS:
        return text
    if text.startswith("90"):
        return "90+"
    return text


def _load_pw_metas(period_start: date | None) -> dict[str, dict]:
    out = {tramo: {"ponderador": None, "meta_total": None} for tramo in PW_CIERRE_TRAMOS}
    if period_start is None:
        return out

    sql = f"""
    SELECT
        tramo,
        CAST(ponderador AS float) AS ponderador,
        CAST(meta_total AS float) AS meta_total
    FROM {PW_METAS_TABLE}
    WHERE periodo = ?
      AND activo = 1
    """
    try:
        rows = run_query(sql, (period_start,))
    except Exception:
        # Si la tabla aun no existe (o hay error de lectura), el cierre debe quedar vacio,
        # no botar todo el dashboard de Porsche.
        return out

    for row in rows:
        tramo = _normalize_pw_tramo(row.get("tramo"))
        if tramo not in out:
            continue
        out[tramo] = {
            "ponderador": row.get("ponderador"),
            "meta_total": row.get("meta_total"),
        }
    return out


def _meta_for(tramo: str, kind: str) -> float:
    tramo_start = _parse_tramo_start(tramo)

    def tramo_high_default(meta_map: dict[str, float]) -> float:
        if tramo_start is not None and tramo_start >= 91:
            # Tramos altos (incluye extras como 241-270) usan meta del bloque alto.
            return meta_map.get("211-240", meta_map.get("181-210", 0.0))
        return 0.0

    if kind == "contactabilidad":
        return CONTACTO_META.get(tramo, tramo_high_default(CONTACTO_META))
    if kind == "recuperacion":
        return RECUPERACION_META.get(tramo, tramo_high_default(RECUPERACION_META))
    if kind == "promesas_pago":
        return 0.60
    if kind == "promesas_cumplidas":
        return 0.80
    if kind == "promesas_incumplidas":
        return 0.20
    if kind == "contenido":
        return 0.40
    if kind == "normalizado":
        return 0.50
    if kind == "renegociacion_solicitudes":
        return 0.20
    if kind == "renegociacion_efectivas":
        return 0.90
    if kind == "tpr":
        return 25.0
    if kind == "reiteracion_contacto":
        return 3.0
    return 0.0


def _to_float(v) -> float:
    try:
        if v is None:
            return 0.0
        return float(v)
    except Exception:
        return 0.0


def _to_int(v) -> int:
    try:
        if v is None:
            return 0
        return int(float(v))
    except Exception:
        return 0


def _parse_tramo_start(tramo: str) -> int | None:
    text = str(tramo or "").strip()
    m = text.split("-", 1)[0].strip()
    if not m.isdigit():
        return None
    return int(m)


def _cuadro_group_for_tramo(tramo: str) -> str | None:
    start = _parse_tramo_start(tramo)
    if start is None:
        return None
    if start == 31:
        return "31-60"
    if start == 61:
        return "61-90"
    if 91 <= start <= 210:
        return "91+"
    return None


def _build_cuadro_contenido(rows: list[dict], metas_by_tramo: dict[str, dict]) -> dict:
    agg = {
        "31-60": {"contenido": 0.0, "no_contenido": 0.0, "total": 0.0},
        "61-90": {"contenido": 0.0, "no_contenido": 0.0, "total": 0.0},
        "91+": {"contenido": 0.0, "no_contenido": 0.0, "total": 0.0},
    }

    for row in rows:
        tramo = str(row.get("tramo", ""))
        group = _cuadro_group_for_tramo(tramo)
        if not group:
            continue
        deuda = _to_float(row.get("deuda"))
        total_pagado = _to_float(row.get("total_pagado"))
        if total_pagado > 0:
            agg[group]["contenido"] += deuda
            agg[group]["total"] += deuda
        else:
            agg[group]["no_contenido"] += deuda
            agg[group]["total"] += deuda

    calc = {}
    group_to_tramo_meta = {"31-60": "31-60", "61-90": "61-90", "91+": "90+"}

    for group in ("31-60", "61-90", "91+"):
        total = agg[group]["total"]
        tramo_meta = group_to_tramo_meta[group]
        meta_data = metas_by_tramo.get(tramo_meta, {})
        meta = meta_data.get("meta_total")
        ponderador = meta_data.get("ponderador")

        if meta is None or ponderador is None:
            calc[group] = {
                "ponderador": None,
                "meta_total": None,
                "real_total": None,
                "cumplimiento": None,
                "resultado": None,
            }
            continue

        real_total = (agg[group]["contenido"] / total) if total else 0.0
        cumplimiento = (real_total / meta) if meta else None
        resultado = (ponderador * cumplimiento) if cumplimiento is not None else None
        calc[group] = {
            "ponderador": ponderador,
            "meta_total": meta,
            "real_total": real_total,
            "cumplimiento": cumplimiento,
            "resultado": resultado,
        }

    rows_visual = [
        {"negocio_pw": "31-60", **calc["31-60"]},
        {"negocio_pw": "61-90", **calc["61-90"]},
        {"negocio_pw": "91-120", "ponderador": None, "meta_total": None, "real_total": None, "cumplimiento": None, "resultado": None},
        {"negocio_pw": "121-150", **calc["91+"]},
        {"negocio_pw": "151-180", "ponderador": None, "meta_total": None, "real_total": None, "cumplimiento": None, "resultado": None},
        {"negocio_pw": "181-210", "ponderador": None, "meta_total": None, "real_total": None, "cumplimiento": None, "resultado": None},
    ]

    resultados = [
        calc["31-60"]["resultado"],
        calc["61-90"]["resultado"],
        calc["91+"]["resultado"],
    ]
    resultados_validos = [value for value in resultados if value is not None]

    return {
        "rows": rows_visual,
        "resultado_total": sum(resultados_validos) if resultados_validos else None,
    }


def _read_dashboard_rows() -> list[dict]:
    return run_query(f"SELECT * FROM {DASHBOARD_VIEW}")


def _available_months(rows: list[dict]) -> list[str]:
    months = {
        _month_label(row.get("fecha_hora_carga"))
        for row in rows
        if row.get("fecha_hora_carga") is not None
    }
    return sorted((m for m in months if m), key=_month_sort_key)


def _tramo_rows(rows: list[dict]) -> list[str]:
    # Orden dinámico desde asignación del mes seleccionado, preservando estándar primero.
    seen = set()
    tramos_data: list[str] = []
    for row in rows:
        tramo = str(row.get("tramo", "")).strip()
        if tramo and tramo not in seen:
            seen.add(tramo)
            tramos_data.append(tramo)

    if not tramos_data:
        return TRAMO_ORDER.copy()

    def sort_key(tramo: str) -> tuple[int, int, str]:
        start = _parse_tramo_start(tramo)
        if start is None:
            return (1, 10_000, tramo)
        return (0, start, tramo)

    # Asegurar tramos estándar en su lugar si existen en datos.
    ordered_standard = [t for t in TRAMO_ORDER if t in seen]
    remaining = [t for t in tramos_data if t not in set(ordered_standard)]
    remaining.sort(key=sort_key)
    return ordered_standard + remaining


def _group_counts(rows: list[dict]) -> dict[str, int]:
    out = defaultdict(int)
    for row in rows:
        tramo = str(row.get("tramo", ""))
        out[tramo] += 1
    return out


def _section_ratio(
    rows: list[dict],
    value_col: str,
    meta_kind: str,
    result_name: str,
    pct_name: str,
    denominator_col: str | None = None,
    incumplida_rule: bool = False,
    tramos: list[str] | None = None,
) -> list[dict]:
    tramos = tramos or _tramo_rows(rows)
    asignado = _group_counts(rows)
    result_sum = defaultdict(float)
    denom_sum = defaultdict(float)

    for row in rows:
        tramo = str(row.get("tramo", ""))
        result_sum[tramo] += _to_float(row.get(value_col))
        if denominator_col:
            denom_sum[tramo] += _to_float(row.get(denominator_col))

    out = []
    for tramo in tramos:
        den = denom_sum[tramo] if denominator_col else float(asignado[tramo])
        res = result_sum[tramo]
        pct = (res / den) if den else 0.0
        meta = _meta_for(tramo, meta_kind)
        if incumplida_rule:
            brecha = meta - pct
            cumplimiento = brecha + 1
        else:
            brecha = pct - meta
            cumplimiento = min(1.0, (pct / meta)) if meta else 0.0
        out.append(
            {
                "tramo": tramo,
                "meta": meta,
                "asignado": int(den if denominator_col else asignado[tramo]),
                result_name: res,
                pct_name: pct,
                "brecha": brecha,
                "cumplimiento": cumplimiento,
            }
        )
    return out


def _section_tpr(rows: list[dict], tramos: list[str] | None = None) -> list[dict]:
    tramos = tramos or _tramo_rows(rows)
    asignado = _group_counts(rows)
    sum_days = defaultdict(float)
    cnt_days = defaultdict(int)
    for row in rows:
        v = row.get("dias_habiles")
        if v is not None and _to_int(row.get("contactabilidad")) == 1:
            tramo = str(row.get("tramo", ""))
            sum_days[tramo] += _to_float(v)
            cnt_days[tramo] += 1

    out = []
    for tramo in tramos:
        tpr = (sum_days[tramo] / cnt_days[tramo]) if cnt_days[tramo] else 0.0
        meta = _meta_for(tramo, "tpr")
        pct_kpi = min(1.0, (meta / tpr)) if tpr else 0.0
        brecha = 0.0 if pct_kpi == 1 else -1 * (pct_kpi + 1)
        out.append(
            {
                "tramo": tramo,
                "meta": meta,
                "asignado": asignado[tramo],
                "TPR": tpr,
                "pct_kpi": pct_kpi,
                "brecha": brecha,
                "cumplimiento": pct_kpi,
            }
        )
    return out


def _section_reiteracion(rows: list[dict], tramos: list[str] | None = None) -> list[dict]:
    tramos = tramos or _tramo_rows(rows)
    asignado = _group_counts(rows)
    contactados = defaultdict(float)
    sum_int = defaultdict(float)
    cnt_int = defaultdict(int)
    for row in rows:
        tramo = str(row.get("tramo", ""))
        c = _to_float(row.get("contactabilidad"))
        contactados[tramo] += c
        if c == 0:
            sum_int[tramo] += _to_float(row.get("intensidad"))
            cnt_int[tramo] += 1

    out = []
    for tramo in tramos:
        rc = (sum_int[tramo] / cnt_int[tramo]) if cnt_int[tramo] else 0.0
        meta = _meta_for(tramo, "reiteracion_contacto")
        pct_kpi = (rc / meta) if meta else 0.0
        brecha = pct_kpi - 1
        cumplimiento = 1.0 if pct_kpi > 1 else pct_kpi
        out.append(
            {
                "tramo": tramo,
                "meta": meta,
                "asignado": asignado[tramo],
                "casos_sin_contacto": int(asignado[tramo] - contactados[tramo]),
                "RC": rc,
                "pct_kpi": pct_kpi,
                "brecha": brecha,
                "cumplimiento": cumplimiento,
            }
        )
    return out


def _section_campana_renegociacion(rows: list[dict]) -> list[dict]:
    # `marca_rene` now carries `identificador_servipag` for export.
    # KPI base must remain binary (has renegotiation marker or not).
    casos_campana = int(
        sum(
            _to_float(
                r.get("marca_rene_bin")
                if r.get("marca_rene_bin") is not None
                else (1 if r.get("marca_rene") not in (None, "", 0, "0") else 0)
            )
            for r in rows
        )
    )
    solicitudes = float(sum(_to_float(r.get("rene_ofrecida")) for r in rows))
    efectivas = float(sum(_to_float(r.get("rene_cursada")) for r in rows))

    def row(nombre: str, meta: float, valor: float) -> dict:
        pct = (valor / casos_campana) if casos_campana else 0.0
        return {
            "tramo": nombre,
            "meta": meta,
            "asignado": casos_campana,
            "kpi": valor,
            "pct_kpi": pct,
            "brecha": pct - meta,
            "cumplimiento": pct,
        }

    return [row("Solicitudes", 0.20, solicitudes), row("Efectivas", 0.90, efectivas)]


def _is_si(value) -> bool:
    return str(value or "").strip().upper() == "SI"


def _build_dashboard_sections(rows: list[dict], tramos: list[str] | None = None) -> dict:
    for r in rows:
        r["contenido_bin"] = 1 if _is_si(r.get("contenido")) else 0
        r["normalizado_bin"] = 1 if _is_si(r.get("normaliza")) else 0
        r["pago_bin"] = 1 if _to_float(r.get("total_pagado")) > 0 else 0

    return {
        "contactabilidad": _section_ratio(rows, "contactabilidad", "contactabilidad", "casos_contactados", "pct_contacto", tramos=tramos),
        "promesas_pago": _section_ratio(rows, "filtro_compromiso", "promesas_pago", "casos_con_promesa", "pct_promesa_pago", denominator_col="contactabilidad", tramos=tramos),
        # IPC/IPI: considerar solo compromisos con gestion valida (CONTACTO DIRECTO o ENVIO WHATSAPP).
        "promesas_cumplidas": _section_ratio(rows, "pago_con_compromiso_kpi", "promesas_cumplidas", "promesas_cumplidas", "pct_cumplimiento_promesa", denominator_col="filtro_compromiso_kpi", tramos=tramos),
        "promesas_incumplidas": _section_ratio(rows, "incumplido_kpi", "promesas_incumplidas", "promesas_incumplidas", "pct_incumplido", denominator_col="filtro_compromiso_kpi", incumplida_rule=True, tramos=tramos),
        "recuperacion": _section_ratio(rows, "pago_bin", "recuperacion", "casos_pagados", "pct_recupero", tramos=tramos),
        "contenido": _section_ratio(rows, "contenido_bin", "contenido", "casos_contenidos", "pct_contenido", tramos=tramos),
        "normalizado": _section_ratio(rows, "normalizado_bin", "normalizado", "casos_normalizados", "pct_normalizado", tramos=tramos),
        "campana_renegociacion": _section_campana_renegociacion(rows),
        "tpr": _section_tpr(rows, tramos=tramos),
        "reiteracion_contacto": _section_reiteracion(rows, tramos=tramos),
    }


def _safe_div(num: float, den: float) -> float:
    return (num / den) if den else 0.0


def _sum_key(rows: list[dict], key: str) -> float:
    return sum(_to_float(row.get(key)) for row in rows)


def _avg_key(rows: list[dict], key: str) -> float:
    values = [_to_float(row.get(key)) for row in rows if row.get(key) not in (None, "")]
    return sum(values) / len(values) if values else 0.0


def _avg_meta(rows: list[dict]) -> float:
    return _avg_key(rows[:4], "meta")


def _build_export_total(section_key: str, rows: list[dict]) -> dict:
    config = EXPORT_DASHBOARD_CONFIG[section_key]
    assigned_key = config["assigned_key"]
    value_key = config["value_key"]
    pct_key = config["pct_key"]

    if section_key == "campana_renegociacion":
        assigned = _to_float(rows[0].get(assigned_key)) if rows else 0.0
        value = _sum_key(rows, value_key)
        meta = _avg_key(rows, "meta")
    elif section_key in ("tpr", "reiteracion_contacto"):
        populated_rows = [row for row in rows if _to_float(row.get(assigned_key))]
        assigned = _sum_key(rows, assigned_key)
        value = _avg_key(populated_rows or rows, value_key)
        meta = _avg_meta(rows)
    else:
        assigned = _sum_key(rows, assigned_key)
        value = _sum_key(rows, value_key)
        meta = _avg_meta(rows)

    pct = _safe_div(value, assigned) if section_key not in ("tpr", "reiteracion_contacto") else _safe_div(value, meta)
    brecha = None if section_key in ("tpr", "reiteracion_contacto") else pct - meta

    return {"tramo": "TOTAL", "meta": meta, assigned_key: assigned, value_key: value, pct_key: pct, "brecha": brecha, "cumplimiento": None}


def _template_candidates() -> list[Path]:
    candidates: list[Path] = []
    env_path = os.getenv(PORSCHE_EXPORT_TEMPLATE_ENV)
    if env_path:
        candidates.append(Path(env_path).expanduser())

    home = Path.home()
    filename = "SEGUIMIENTO_PORSCHE_20260527.xlsx"
    candidates.extend(
        [
            home / "OneDrive - Phoenix Service" / "Escritorio" / "SEGUIMIENTOS" / "PORSCHE" / "Mayo 2026" / filename,
            home / "Desktop" / "SEGUIMIENTOS" / "PORSCHE" / "Mayo 2026" / filename,
            home / "OneDrive - Phoenix Service" / "Desktop" / "SEGUIMIENTOS" / "PORSCHE" / "Mayo 2026" / filename,
        ]
    )
    return candidates


def _load_export_template():
    attempted: list[str] = []
    for path in _template_candidates():
        if not path.exists():
            attempted.append(f"{path} (no existe)")
            continue
        try:
            return load_workbook(path)
        except Exception as exc:
            attempted.append(f"{path} ({type(exc).__name__}: {exc})")

    raise FileNotFoundError(f"No se pudo abrir la plantilla Porsche. Rutas revisadas: {'; '.join(attempted)}")


def _copy_cell_style(source, target) -> None:
    if not source.has_style:
        return
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy(source.protection)


def _copy_row_format(ws, source_row: int, target_row: int, max_col: int) -> None:
    if target_row == source_row:
        return
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, max_col + 1):
        _copy_cell_style(ws.cell(source_row, col), ws.cell(target_row, col))


def _ensure_rows(ws, max_row: int, template_row: int, max_col: int) -> None:
    for row_idx in range(ws.max_row + 1, max_row + 1):
        ws.cell(row_idx, max_col).value = None
    for row_idx in range(template_row, max_row + 1):
        _copy_row_format(ws, template_row, row_idx, max_col)


def _clear_values(ws, min_row: int = 1) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _lookup(row: dict, keys: tuple[str, ...], default=None):
    lowered = {str(k).lower(): v for k, v in row.items()}
    for key in keys:
        if key in row:
            return row[key]
        lowered_key = key.lower()
        if lowered_key in lowered:
            return lowered[lowered_key]
    return default


def _rut_number(value):
    if value is None:
        return None
    text = str(value).strip().replace(".", "")
    number = text.split("-", 1)[0].strip()
    return int(number) if number.isdigit() else value


def _si_no(value):
    if value in (None, ""):
        return value
    text = str(value).strip().upper()
    if text in {"1", "SI", "S", "TRUE"}:
        return "si"
    if text in {"0", "NO", "N", "FALSE"}:
        return "no"
    return value


def _data_export_values(row: dict) -> list:
    values = []
    for idx, keys in enumerate(DATA_EXPORT_COLUMNS):
        value = _lookup(row, keys)
        if idx == 2:
            value = _rut_number(value)
        elif idx in (12, 13):
            value = _si_no(value)
        elif idx == 16 and value in (None, ""):
            value = 1 if _to_float(_lookup(row, ("total_pagado", "pagos", "monto_pagado"))) > 0 else 0
        elif idx == 17 and value in (None, ""):
            value = 0 if _to_float(_lookup(row, ("total_pagado", "pagos", "monto_pagado"))) > 0 else 1
        values.append(value)
    return values


def _write_data_sheet(wb, rows: list[dict]) -> None:
    if "DATA" not in wb.sheetnames:
        return

    ws = wb["DATA"]
    max_col = max(ws.max_column, len(DATA_EXPORT_COLUMNS))
    last_row = max(3, len(rows) + 2)
    _ensure_rows(ws, last_row, 3, max_col)
    _clear_values(ws, min_row=3)

    for idx, row in enumerate(rows, start=3):
        _copy_row_format(ws, 3, idx, max_col)
        for col, value in enumerate(_data_export_values(row), start=1):
            ws.cell(idx, col).value = value


def _write_dashboard_row(ws, row_idx: int, row: dict, config: dict) -> None:
    assigned_key = config["assigned_key"]
    value_key = config["value_key"]
    pct_key = config["pct_key"]
    values_by_col = {
        "B": row.get("tramo"),
        "D": row.get("meta"),
        "E": row.get(assigned_key),
        "F": row.get(value_key),
        "H": row.get(pct_key),
        "I": row.get("brecha"),
        "J": row.get("cumplimiento"),
    }
    for col, value in values_by_col.items():
        ws[f"{col}{row_idx}"].value = value


def _write_dashboard_sheet(wb, sections: dict) -> None:
    if "Dashboard" not in wb.sheetnames:
        return

    ws = wb["Dashboard"]
    for section_key, config in EXPORT_DASHBOARD_CONFIG.items():
        labels = config.get("labels", EXPORT_TRAMO_ORDER)
        rows_by_label = {str(row.get("tramo")): row for row in sections.get(section_key, [])}
        display_rows = [rows_by_label.get(label, {"tramo": label}) for label in labels]

        for idx, row in enumerate(display_rows, start=config["start_row"]):
            _write_dashboard_row(ws, idx, row, config)

        _write_dashboard_row(ws, config["total_row"], _build_export_total(section_key, display_rows), config)


def _keep_export_sheets(wb) -> None:
    missing = [sheet_name for sheet_name in EXPORT_VISIBLE_SHEETS if sheet_name not in wb.sheetnames]
    if missing:
        raise ValueError(f"La plantilla Porsche no tiene las hojas requeridas: {', '.join(missing)}")

    for sheet_name in list(wb.sheetnames):
        if sheet_name not in EXPORT_VISIBLE_SHEETS:
            del wb[sheet_name]

    wb.active = wb.sheetnames.index("Dashboard")


def _build_porsche_export_workbook(rows: list[dict]) -> BytesIO:
    wb = _load_export_template()
    sections = _build_dashboard_sections(rows, tramos=EXPORT_TRAMO_ORDER)
    _write_data_sheet(wb, rows)
    _write_dashboard_sheet(wb, sections)
    _keep_export_sheets(wb)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _filename_date(selected_month: str) -> str:
    period_start = _period_start_from_month_label(selected_month)
    if period_start is None:
        return datetime.now().strftime("%Y%m%d")
    return f"{period_start.year}{period_start.month:02d}{datetime.now().day:02d}"


@router.get("/filtros")
def filtros() -> dict:
    try:
        rows = _read_dashboard_rows()
        meses = _available_months(rows)
        return {"filters": {"meses": meses, "default_mes": meses[-1] if meses else ""}}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@router.get("/dashboard")
def dashboard(mes: str | None = None) -> dict:
    try:
        rows = _read_dashboard_rows()
        available_months = _available_months(rows)
        selected_month = mes or (available_months[-1] if available_months else "")
        rows = _selected_month_rows(rows, selected_month)
        sections = _build_dashboard_sections(rows)

        summary = {
            "asignados": len(rows),
            "contactados": int(sum(_to_float(r.get("contactabilidad")) for r in rows)),
            "con_promesa": int(sum(_to_float(r.get("filtro_compromiso")) for r in rows)),
            "recuperado": float(sum(_to_float(r.get("total_pagado")) for r in rows)),
        }

        return {
            "filters": {"meses": available_months, "default_mes": available_months[-1] if available_months else "", "mes": selected_month},
            "summary": summary,
            "sections": sections,
        }
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@router.get("/cuadro-contenido")
def cuadro_contenido(mes: str | None = None) -> dict:
    try:
        rows = _read_dashboard_rows()
        available_months = _available_months(rows)
        selected_month = mes or (available_months[-1] if available_months else "")
        month_rows = _selected_month_rows(rows, selected_month)
        period_start = _period_start_from_month_label(selected_month)
        metas_by_tramo = _load_pw_metas(period_start)
        cuadro = _build_cuadro_contenido(month_rows, metas_by_tramo)
        return {
            "filters": {"meses": available_months, "default_mes": available_months[-1] if available_months else "", "mes": selected_month},
            "cuadro": cuadro,
        }
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@router.get("/porsche/export")
def export_porsche(mes: str | None = Query(default=None)) -> StreamingResponse:
    try:
        rows = _read_dashboard_rows()
        available_months = _available_months(rows)
        selected_month = mes or (available_months[-1] if available_months else "")
        month_rows = _selected_month_rows(rows, selected_month)
        output = _build_porsche_export_workbook(month_rows)
        filename = f"SEGUIMIENTO_PORSCHE_{_filename_date(selected_month)}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
