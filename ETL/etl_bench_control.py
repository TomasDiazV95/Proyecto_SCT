import argparse
import base64
import json
import logging
import os
import re
import sys
import tempfile
import unicodedata
from contextlib import suppress
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
import pyodbc
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook


TABLE = "dbo.tmp_BENCH_CONTROL_DIARIO"
DEFAULT_LOG_DIR = Path(__file__).resolve().parents[1] / "Logs"
STATE_FILE = DEFAULT_LOG_DIR / "etl_bench_control_state.json"
SCHEMA_VERSION = "2026-08-31-business-days-remaining"
BENCH_FILENAME = "Control semanal puesto BENCH.xlsx"
DEFAULT_GRAPH_SHARE_URL = (
    "https://phoenixservice1.sharepoint.com/:x:/r/sites/Cobranzas/_layouts/15/Doc.aspx?"
    "sourcedoc=%7B9743AD98-27D5-40ED-BEE6-919B57F3D447%7D&file=Control%20semanal%20puesto%20BENCH.xlsx"
    "&action=default&mobileredirect=true"
)
GRAPH_API_BASE = "https://graph.microsoft.com/v1.0"
HEADER_EMPRESA = "EMPRESA"
HEADER_SEGMENTO = "SEGMENTO"
SPANISH_MONTHS = {
    "ENERO",
    "FEBRERO",
    "MARZO",
    "ABRIL",
    "MAYO",
    "JUNIO",
    "JULIO",
    "AGOSTO",
    "SEPTIEMBRE",
    "OCTUBRE",
    "NOVIEMBRE",
    "DICIEMBRE",
}

# Keep this catalog explicit so changes to Chilean holidays are reviewable.
CHILE_HOLIDAYS_BY_YEAR: dict[int, frozenset[date]] = {
    2026: frozenset(
        {
            date(2026, 1, 1),
            date(2026, 4, 3),
            date(2026, 4, 4),
            date(2026, 5, 1),
            date(2026, 5, 21),
            date(2026, 6, 21),
            date(2026, 6, 29),
            date(2026, 7, 16),
            date(2026, 8, 15),
            date(2026, 9, 18),
            date(2026, 9, 19),
            date(2026, 10, 12),
            date(2026, 10, 31),
            date(2026, 11, 1),
            date(2026, 12, 8),
            date(2026, 12, 25),
        }
    ),
}


def load_env_files() -> None:
    root_dir = Path(__file__).resolve().parents[1]
    load_dotenv(root_dir / ".env")
    load_dotenv(root_dir / "ETL" / ".env")


def setup_logging() -> Path:
    log_dir = Path(os.getenv("BENCH_LOG_DIR") or DEFAULT_LOG_DIR)
    log_dir.mkdir(parents=True, exist_ok=True)
    log_path = log_dir / f"etl_bench_control_{datetime.now().strftime('%Y%m%d')}.log"
    file_handler = logging.FileHandler(log_path, encoding="utf-8")
    stream_handler = logging.StreamHandler()
    file_handler.setLevel(logging.INFO)
    stream_handler.setLevel(logging.ERROR)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        handlers=[file_handler, stream_handler],
    )
    return log_path


def emit_payload(payload: dict[str, Any]) -> None:
    print(json.dumps(payload, ensure_ascii=False, default=str))


def read_state_file() -> dict[str, Any]:
    if not STATE_FILE.exists():
        return {}
    try:
        return json.loads(STATE_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}


def write_state_file(state: dict[str, Any]) -> None:
    STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
    STATE_FILE.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


def _env_first(*names: str) -> str:
    for name in names:
        value = os.getenv(name, "").strip()
        if value:
            return value
    return ""


def _graph_access_token() -> str:
    tenant_id = _env_first("GRAPH_TENANT_ID", "AZURE_TENANT_ID")
    client_id = _env_first("GRAPH_CLIENT_ID", "AZURE_CLIENT_ID")
    client_secret = _env_first("GRAPH_CLIENT_SECRET", "AZURE_CLIENT_SECRET")

    if tenant_id and client_id and client_secret:
        token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
        form = {
            "client_id": client_id,
            "client_secret": client_secret,
            "scope": "https://graph.microsoft.com/.default",
            "grant_type": "client_credentials",
        }
        token_response = requests.post(token_url, data=form, timeout=30)
        if token_response.status_code != 200:
            raise RuntimeError(
                f"No se pudo obtener token Graph: {token_response.status_code} {token_response.text}"
            )
        body = token_response.json()
        access_token = str(body.get("access_token") or "").strip()
        if not access_token:
            raise RuntimeError("Azure no devolvio access_token para Graph")
        return access_token

    fallback = _env_first("GRAPH_ACCESS_TOKEN")
    if fallback:
        return fallback

    raise RuntimeError(
        "Faltan credenciales de Graph. Define AZURE_TENANT_ID/AZURE_CLIENT_ID/AZURE_CLIENT_SECRET o GRAPH_ACCESS_TOKEN"
    )


def _graph_headers() -> dict[str, str]:
    return {
        "Authorization": f"Bearer {_graph_access_token()}",
        "Accept": "application/json",
    }


def _graph_share_url() -> str:
    return (os.getenv("BENCH_GRAPH_SHARE_URL") or DEFAULT_GRAPH_SHARE_URL).strip()


def _encode_share_url(share_url: str) -> str:
    encoded = base64.b64encode(share_url.encode("utf-8")).decode("ascii")
    encoded = encoded.rstrip("=").replace("+", "-").replace("/", "_")
    return f"u!{encoded}"


def _graph_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    cleaned = value.strip()
    if cleaned.endswith("Z"):
        cleaned = cleaned[:-1] + "+00:00"
    try:
        return datetime.fromisoformat(cleaned)
    except ValueError:
        return None


def _graph_bench_metadata() -> dict[str, Any]:
    share_url = _graph_share_url()
    endpoint = f"{GRAPH_API_BASE}/shares/{_encode_share_url(share_url)}/driveItem"
    response = requests.get(
        endpoint,
        headers=_graph_headers(),
        params={"$select": "id,name,eTag,lastModifiedDateTime,@microsoft.graph.downloadUrl"},
        timeout=60,
    )
    if response.status_code != 200:
        raise RuntimeError(f"No se pudo leer el archivo BENCH desde Graph: {response.status_code} {response.text}")

    payload = response.json()
    download_url = str(payload.get("@microsoft.graph.downloadUrl") or "").strip()
    if not download_url:
        download_url = f"{GRAPH_API_BASE}/shares/{_encode_share_url(share_url)}/driveItem/content"

    return {
        "share_url": share_url,
        "item_id": str(payload.get("id") or "").strip(),
        "file_name": str(payload.get("name") or BENCH_FILENAME).strip() or BENCH_FILENAME,
        "etag": str(payload.get("eTag") or "").strip(),
        "last_modified": _graph_datetime(str(payload.get("lastModifiedDateTime") or "").strip()),
        "download_url": download_url,
    }


def _download_graph_bench_file(
    target_dir: Path | None = None,
    metadata: dict[str, Any] | None = None,
) -> tuple[Path, dict[str, Any]]:
    metadata = metadata or _graph_bench_metadata()
    target_dir = target_dir or Path(tempfile.gettempdir())
    target_dir.mkdir(parents=True, exist_ok=True)

    safe_name = re.sub(r"[^A-Za-z0-9._\- ]", "_", metadata["file_name"])
    etag_suffix = re.sub(r"[^A-Za-z0-9]+", "_", metadata["etag"] or "noetag").strip("_") or "noetag"
    download_path = target_dir / f"bench_control_{etag_suffix}_{safe_name}"

    download_url = str(metadata["download_url"])
    if download_url.startswith(GRAPH_API_BASE):
        response = requests.get(download_url, headers=_graph_headers(), timeout=120, allow_redirects=True)
    else:
        response = requests.get(download_url, timeout=120)
    if response.status_code != 200:
        raise RuntimeError(f"No se pudo descargar el archivo BENCH desde Graph: {response.status_code} {response.text}")

    download_path.write_bytes(response.content)
    logging.info("Archivo BENCH descargado desde Graph: %s", download_path)
    return download_path, metadata


def normalize_text(value: object) -> str:
    text = str(value or "").strip()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("utf-8")
    return " ".join(text.upper().split())


def pick_driver() -> str:
    available = list(pyodbc.drivers())
    preferred = [
        os.getenv("DB_DRIVER"),
        "ODBC Driver 18 for SQL Server",
        "ODBC Driver 17 for SQL Server",
        "SQL Server",
    ]
    for driver in preferred:
        if driver and driver in available:
            return driver
    raise RuntimeError(f"No hay driver ODBC para SQL Server. Drivers encontrados: {available}")


def encrypt_candidates() -> list[str | None]:
    encrypt_env = os.getenv("DB_ENCRYPT")
    if encrypt_env is not None and str(encrypt_env).strip():
        return [str(encrypt_env).strip()]
    return ["yes", "no", None]


def connect() -> pyodbc.Connection:
    server = os.getenv("DB_SERVER")
    database = os.getenv("DB_NAME")
    user = os.getenv("DB_USER")
    password = os.getenv("DB_PASSWORD")
    if not all([server, database, user, password]):
        raise RuntimeError("Faltan DB_SERVER, DB_NAME, DB_USER o DB_PASSWORD")

    drivers_to_try: list[str] = []
    preferred = os.getenv("DB_DRIVER")
    if preferred:
        drivers_to_try.append(preferred)
    for candidate in ("ODBC Driver 18 for SQL Server", "ODBC Driver 17 for SQL Server", "SQL Server"):
        if candidate not in drivers_to_try:
            drivers_to_try.append(candidate)

    available = set(pyodbc.drivers())
    errors: list[str] = []
    for candidate_driver in drivers_to_try:
        if candidate_driver not in available:
            continue
        for encrypt in encrypt_candidates():
            parts = [
                f"Driver={{{candidate_driver}}}",
                f"Server={server}",
                f"Database={database}",
                f"Uid={user}",
                f"Pwd={password}",
                "TrustServerCertificate=yes",
            ]
            if candidate_driver != "SQL Server" and encrypt is not None:
                parts.append(f"Encrypt={encrypt}")
            conn_str = ";".join(parts) + ";"
            try:
                return pyodbc.connect(conn_str)
            except pyodbc.Error as exc:
                errors.append(f"{candidate_driver} / Encrypt={encrypt}: {exc}")

    raise RuntimeError("No se pudo conectar a SQL Server. Intentos: " + " | ".join(errors))


def resolve_excel_path(file_path: str | None) -> Path:
    raise RuntimeError("La fuente BENCH ahora es fija por Microsoft Graph y no usa rutas locales.")


def discover_bench_file() -> Path | None:
    search_roots: list[Path] = []
    user_home = Path.home()
    phoenix_service_root = user_home / "Phoenix Service"
    if phoenix_service_root.exists() and phoenix_service_root.is_dir():
        search_roots.append(phoenix_service_root)
    for candidate in user_home.glob("OneDrive*"):
        if candidate.exists() and candidate.is_dir():
            search_roots.append(candidate)

    matches: list[Path] = []
    for root in search_roots:
        with suppress(OSError, PermissionError):
            matches.extend(root.rglob(BENCH_FILENAME))

    files = [match for match in matches if match.is_file()]
    if not files:
        return None
    files.sort(key=lambda item: item.stat().st_mtime, reverse=True)
    return files[0]


def refresh_workbook_cache(path: Path) -> None:
    should_refresh = str(os.getenv("BENCH_REFRESH_WITH_EXCEL", "1")).strip().lower()
    if should_refresh in {"0", "false", "no"}:
        return
    if os.name != "nt":
        return

    try:
        import pythoncom
        import win32com.client
    except Exception as exc:
        logging.warning("No se pudo habilitar recálculo con Excel para %s: %s", path, exc)
        return

    excel = None
    workbook = None
    try:
        pythoncom.CoInitialize()
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        workbook = excel.Workbooks.Open(str(path), UpdateLinks=0, ReadOnly=False)
        excel.CalculateFullRebuild()
        workbook.Save()
        logging.info("Excel recalc/guardado aplicado a %s", path)
    except Exception as exc:
        logging.warning("No se pudo refrescar el workbook con Excel para %s: %s", path, exc)
    finally:
        with suppress(Exception):
            if workbook is not None:
                workbook.Close(SaveChanges=False)
        with suppress(Exception):
            if excel is not None:
                excel.Quit()
        with suppress(Exception):
            pythoncom.CoUninitialize()


def clean_string(value: object) -> str | None:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    if text.lower() in {"nan", "nat", "none"}:
        return None
    return text or None


def parse_date(value: object) -> date | None:
    if value is None:
        return None
    parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        return None
    return parsed.date()


def parse_percentage(value: object, number_format: str | None = None) -> float | None:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, bool):
        return None

    fmt = str(number_format or "")
    if isinstance(value, (int, float)):
        number = float(value)
        if "%" in fmt:
            number *= 100
        return round(number, 4)

    original_text = str(value).strip()
    if not original_text:
        return None

    has_percent = "%" in original_text or "%" in fmt
    text = original_text.replace("%", "").replace(" ", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(".", "").replace(",", ".")

    try:
        number = float(text)
    except ValueError:
        return None

    if has_percent and "%" not in original_text and -1 <= number <= 1:
        number *= 100
    return round(number, 4)


def parse_business_day(value: object) -> int | None:
    cleaned = clean_string(value)
    if cleaned is None:
        return None
    try:
        return int(float(str(cleaned).replace(",", ".")))
    except (ValueError, OverflowError):
        return None


def chile_holidays_for_year(year: int) -> frozenset[date]:
    try:
        return CHILE_HOLIDAYS_BY_YEAR[year]
    except KeyError as exc:
        available_years = ", ".join(str(item) for item in sorted(CHILE_HOLIDAYS_BY_YEAR))
        raise RuntimeError(
            f"No hay catalogo de feriados de Chile para el anio {year}. "
            f"Anos disponibles: {available_years}"
        ) from exc


def business_days_remaining(fecha: date) -> int:
    """Count business days strictly after fecha until the end of its month."""
    holidays = chile_holidays_for_year(fecha.year)
    next_day = fecha + timedelta(days=1)
    if next_day.month != fecha.month:
        return 0

    month_end = (
        date(fecha.year, fecha.month + 1, 1) - timedelta(days=1)
        if fecha.month < 12
        else date(fecha.year, 12, 31)
    )
    business_days = 0
    current = next_day
    while current <= month_end:
        if current.weekday() < 5 and current not in holidays:
            business_days += 1
        current += timedelta(days=1)
    return business_days


def safe_period_values(df: pd.DataFrame) -> list[str]:
    if df.empty or "fecha" not in df.columns:
        return []
    return sorted(df["fecha"].astype(str).str.slice(0, 7).dropna().unique().tolist())


def nullable_float(value: object) -> float | None:
    if value is None or pd.isna(value):
        return None
    return float(value)


def nullable_int(value: object) -> int | None:
    if value is None or pd.isna(value):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError, OverflowError):
        return None


def spanish_month_name(month: int) -> str:
    names = {
        1: "Enero",
        2: "Febrero",
        3: "Marzo",
        4: "Abril",
        5: "Mayo",
        6: "Junio",
        7: "Julio",
        8: "Agosto",
        9: "Septiembre",
        10: "Octubre",
        11: "Noviembre",
        12: "Diciembre",
    }
    return names[month]


def build_merged_value_map(worksheet) -> dict[tuple[int, int], Any]:
    merged_values: dict[tuple[int, int], Any] = {}
    for merged_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_value = worksheet.cell(row=min_row, column=min_col).value
        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                merged_values[(row_idx, col_idx)] = top_left_value
    return merged_values


def worksheet_cell_value(worksheet, merged_values: dict[tuple[int, int], Any], row_idx: int, col_idx: int) -> Any:
    value = worksheet.cell(row=row_idx, column=col_idx).value
    if value is not None:
        return value
    return merged_values.get((row_idx, col_idx))


def is_header_row(values: list[Any]) -> bool:
    if len(values) < 2:
        return False
    return normalize_text(values[0]) == HEADER_EMPRESA and normalize_text(values[1]) == HEADER_SEGMENTO


def is_month_banner_row(values: list[Any]) -> bool:
    if len(values) < 2:
        return False
    second_cell = normalize_text(values[1])
    if second_cell not in SPANISH_MONTHS:
        return False
    numeric_markers = 0
    for value in values[2:]:
        parsed = parse_business_day(value)
        if parsed is not None:
            numeric_markers += 1
    return numeric_markers > 0


def extract_sheet_records(worksheet, sheet_name: str) -> list[dict[str, Any]]:
    merged_values = build_merged_value_map(worksheet)
    records: list[dict[str, Any]] = []
    current_dates: list[tuple[int, date]] = []

    for row_idx in range(1, worksheet.max_row + 1):
        row_values = [
            worksheet_cell_value(worksheet, merged_values, row_idx, col_idx)
            for col_idx in range(1, worksheet.max_column + 1)
        ]

        if is_month_banner_row(row_values):
            continue

        if is_header_row(row_values):
            current_dates = []
            for col_idx in range(3, worksheet.max_column + 1):
                fecha = parse_date(worksheet_cell_value(worksheet, merged_values, row_idx, col_idx))
                if fecha:
                    current_dates.append((col_idx, fecha))
            continue

        if not current_dates:
            continue

        empresa = clean_string(worksheet_cell_value(worksheet, merged_values, row_idx, 1))
        segmento = clean_string(worksheet_cell_value(worksheet, merged_values, row_idx, 2))
        if not empresa or not segmento:
            continue
        if normalize_text(segmento) in SPANISH_MONTHS:
            continue

        for col_idx, fecha in current_dates:
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cumplimiento = parse_percentage(
                worksheet_cell_value(worksheet, merged_values, row_idx, col_idx),
                number_format=str(getattr(cell, "number_format", "") or ""),
            )
            records.append(
                {
                    "negocio": sheet_name,
                    "empresa": empresa,
                    "segmento": segmento,
                    "fecha": fecha,
                    "cumplimiento": cumplimiento,
                }
            )

    return records


def read_excel_source(path: Path, requested_sheet: str | None = None) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    requested_sheet = requested_sheet or os.getenv("BENCH_SHEET_NAME")
    refresh_workbook_cache(path)
    workbook = load_workbook(path, data_only=True)
    normalized_sheets = {normalize_text(name): name for name in workbook.sheetnames}
    selected_sheets: list[str]
    if requested_sheet:
        sheet_name = normalized_sheets.get(normalize_text(requested_sheet))
        if not sheet_name:
            available = ", ".join(workbook.sheetnames)
            raise RuntimeError(f"No existe la hoja '{requested_sheet}' en {path.name}. Hojas disponibles: {available}")
        selected_sheets = [sheet_name]
    else:
        selected_sheets = list(workbook.sheetnames)

    extracted_rows: list[dict[str, Any]] = []
    for sheet_name in selected_sheets:
        extracted_rows.extend(extract_sheet_records(workbook[sheet_name], sheet_name))

    if not extracted_rows:
        raise RuntimeError(f"No se encontraron hojas procesables en {path.name}")

    return extracted_rows, selected_sheets, workbook.sheetnames


def transform_dataframe(extracted_rows: list[dict[str, Any]]) -> tuple[pd.DataFrame, dict[str, int]]:
    records: list[dict[str, object]] = []
    stats = {
        "rows_read": len(extracted_rows),
        "rows_valid": 0,
        "rows_discarded": 0,
        "invalid_fecha": 0,
        "invalid_cumplimiento": 0,
        "missing_cumplimiento": 0,
    }
    fecha_actualizacion = datetime.now().replace(microsecond=0)

    for row in extracted_rows:
        fecha = parse_date(row.get("fecha"))
        negocio = clean_string(row.get("negocio"))
        segmento = clean_string(row.get("segmento"))
        empresa = clean_string(row.get("empresa"))
        cumplimiento = parse_percentage(row.get("cumplimiento"))

        if row.get("fecha") is not None and fecha is None:
            stats["invalid_fecha"] += 1
        if row.get("cumplimiento") is not None and cumplimiento is None:
            stats["invalid_cumplimiento"] += 1

        if not all([fecha, negocio, segmento, empresa]):
            stats["rows_discarded"] += 1
            continue

        if cumplimiento is None:
            stats["missing_cumplimiento"] += 1
            stats["rows_discarded"] += 1
            continue

        records.append(
            {
                "fecha": fecha,
                "anio": fecha.year,
                "mes": spanish_month_name(fecha.month),
                "dia_habil": business_days_remaining(fecha),
                "negocio": negocio,
                "segmento": segmento,
                "empresa": empresa,
                "cumplimiento": cumplimiento,
                "fecha_actualizacion": fecha_actualizacion,
            }
        )

    stats["rows_valid"] = len(records)
    if not records:
        raise RuntimeError("No se encontraron filas validas en el archivo BENCH")
    df = pd.DataFrame.from_records(records)
    df = df.sort_values(
        by=["fecha", "negocio", "segmento", "empresa"],
        ascending=[True, True, True, True],
    )
    df = df.drop_duplicates(subset=["fecha", "negocio", "segmento", "empresa"], keep="first")
    stats["rows_valid"] = len(df.index)
    return df, stats


def get_table_columns(cur: pyodbc.Cursor, table_name: str) -> list[tuple[str, str, str, int | None, int | None]]:
    cur.execute(
        """
        SELECT
            COLUMN_NAME,
            DATA_TYPE,
            IS_NULLABLE,
            NUMERIC_PRECISION,
            NUMERIC_SCALE,
            CHARACTER_MAXIMUM_LENGTH,
            DATETIME_PRECISION
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = 'dbo' AND TABLE_NAME = ?
        ORDER BY ORDINAL_POSITION
        """,
        (table_name.split(".")[-1],),
    )
    signature: list[tuple[str, str, str, int | None, int | None]] = []
    for row in cur.fetchall():
        column_name = str(row[0]).lower()
        data_type = str(row[1]).lower()
        is_nullable = str(row[2]).upper()
        numeric_precision = int(row[3]) if row[3] is not None else None
        numeric_scale = int(row[4]) if row[4] is not None else None
        char_length = int(row[5]) if row[5] is not None else None
        datetime_precision = int(row[6]) if row[6] is not None else None
        size_hint = char_length if char_length is not None else numeric_precision
        if data_type == "datetime2":
            size_hint = datetime_precision
        signature.append((column_name, data_type, is_nullable, size_hint, numeric_scale))
    return signature


def create_bench_table(cur: pyodbc.Cursor) -> None:
    cur.execute(
        f"""
        CREATE TABLE {TABLE} (
            fecha DATE NOT NULL,
            anio INT NOT NULL,
            mes NVARCHAR(20) NOT NULL,
            dia_habil INT NULL,
            negocio NVARCHAR(150) NOT NULL,
            segmento NVARCHAR(150) NOT NULL,
            empresa NVARCHAR(150) NOT NULL,
            cumplimiento DECIMAL(9,4) NULL,
            fecha_actualizacion DATETIME2(0) NOT NULL
        )
        """
    )
    cur.execute(f"CREATE INDEX IX_tmp_BENCH_CONTROL_DIARIO_fecha ON {TABLE}(fecha)")
    cur.execute(f"CREATE INDEX IX_tmp_BENCH_CONTROL_DIARIO_lookup ON {TABLE}(negocio, segmento, empresa, fecha)")
    cur.execute(
        f"""
        CREATE UNIQUE INDEX UX_tmp_BENCH_CONTROL_DIARIO_key
        ON {TABLE}(fecha, negocio, segmento, empresa)
        """
    )


def ensure_table(cur: pyodbc.Cursor) -> None:
    expected_columns = [
        ("fecha", "date", "NO", None, None),
        ("anio", "int", "NO", 10, 0),
        ("mes", "nvarchar", "NO", 20, None),
        ("dia_habil", "int", "YES", 10, 0),
        ("negocio", "nvarchar", "NO", 150, None),
        ("segmento", "nvarchar", "NO", 150, None),
        ("empresa", "nvarchar", "NO", 150, None),
        ("cumplimiento", "decimal", "YES", 9, 4),
        ("fecha_actualizacion", "datetime2", "NO", 0, None),
    ]
    cur.execute(f"SELECT OBJECT_ID('{TABLE}', 'U')")
    exists = cur.fetchone()[0] is not None
    if not exists:
        create_bench_table(cur)
        return

    existing_columns = get_table_columns(cur, TABLE)
    if existing_columns != expected_columns:
        cur.execute(f"DROP TABLE {TABLE}")
        create_bench_table(cur)


def load_already_processed(graph_item_id: str, graph_etag: str) -> bool:
    state = read_state_file()
    return (
        str(state.get("schema_version") or "") == SCHEMA_VERSION
        and str(state.get("graph_item_id") or "") == graph_item_id
        and str(state.get("graph_etag") or "") == graph_etag
    )


def bench_table_has_rows() -> bool:
    """Do not skip a source reload when the destination table is empty."""
    with connect() as cn:
        cur = cn.cursor()
        cur.execute(f"SELECT OBJECT_ID('{TABLE}', 'U')")
        if cur.fetchone()[0] is None:
            return False
        cur.execute(f"SELECT TOP 1 1 FROM {TABLE}")
        return cur.fetchone() is not None


def merge_rows(cur: pyodbc.Cursor, df: pd.DataFrame) -> tuple[int, int, int]:
    cur.execute(
        """
        CREATE TABLE #bench_stage (
            fecha DATE NOT NULL,
            anio INT NOT NULL,
            mes NVARCHAR(20) NOT NULL,
            dia_habil INT NULL,
            negocio NVARCHAR(150) NOT NULL,
            segmento NVARCHAR(150) NOT NULL,
            empresa NVARCHAR(150) NOT NULL,
            cumplimiento DECIMAL(9,4) NULL,
            fecha_actualizacion DATETIME2(0) NOT NULL
        )
        """
    )

    stage_rows = [
        (
            row["fecha"],
            int(row["anio"]),
            row["mes"],
            nullable_int(row.get("dia_habil")),
            row["negocio"],
            row["segmento"],
            row["empresa"],
            nullable_float(row.get("cumplimiento")),
            row["fecha_actualizacion"],
        )
        for _, row in df.iterrows()
    ]
    cur.executemany(
        """
        INSERT INTO #bench_stage (
            fecha, anio, mes, dia_habil, negocio, segmento, empresa, cumplimiento, fecha_actualizacion
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        stage_rows,
    )

    cur.execute(
        """
        CREATE TABLE #bench_scope (
            negocio NVARCHAR(150) NOT NULL,
            periodo CHAR(7) NOT NULL,
            PRIMARY KEY (negocio, periodo)
        )
        """
    )
    scope_rows = sorted(
        {
            (
                str(row["negocio"]),
                str(row["fecha"])[:7],
            )
            for _, row in df.iterrows()
        }
    )
    cur.executemany(
        """
        INSERT INTO #bench_scope (negocio, periodo)
        VALUES (?, ?)
        """,
        scope_rows,
    )

    cur.execute("CREATE TABLE #bench_merge_actions (action_name NVARCHAR(10) NOT NULL)")
    cur.execute(
        f"""
        MERGE {TABLE} AS target
        USING #bench_stage AS source
          ON target.fecha = source.fecha
         AND target.negocio = source.negocio
         AND target.segmento = source.segmento
         AND target.empresa = source.empresa
        WHEN MATCHED AND (
            target.anio <> source.anio
            OR target.mes <> source.mes
            OR ISNULL(target.dia_habil, -1) <> ISNULL(source.dia_habil, -1)
            OR (
                (target.cumplimiento IS NULL AND source.cumplimiento IS NOT NULL)
                OR (target.cumplimiento IS NOT NULL AND source.cumplimiento IS NULL)
                OR (target.cumplimiento IS NOT NULL AND source.cumplimiento IS NOT NULL AND target.cumplimiento <> source.cumplimiento)
            )
        ) THEN
            UPDATE SET
                anio = source.anio,
                mes = source.mes,
                dia_habil = source.dia_habil,
                cumplimiento = source.cumplimiento,
                fecha_actualizacion = source.fecha_actualizacion
        WHEN NOT MATCHED BY TARGET THEN
            INSERT (fecha, anio, mes, dia_habil, negocio, segmento, empresa, cumplimiento, fecha_actualizacion)
            VALUES (
                source.fecha,
                source.anio,
                source.mes,
                source.dia_habil,
                source.negocio,
                source.segmento,
                source.empresa,
                source.cumplimiento,
                source.fecha_actualizacion
            )
        WHEN NOT MATCHED BY SOURCE
         AND EXISTS (
            SELECT 1
            FROM #bench_scope scope
            WHERE scope.negocio = target.negocio
              AND scope.periodo = CONVERT(char(7), target.fecha, 126)
         ) THEN
            DELETE
        OUTPUT $action INTO #bench_merge_actions(action_name);
        """
    )
    cur.execute(
        """
        SELECT
            SUM(CASE WHEN action_name = 'INSERT' THEN 1 ELSE 0 END) AS inserted_rows,
            SUM(CASE WHEN action_name = 'UPDATE' THEN 1 ELSE 0 END) AS updated_rows,
            SUM(CASE WHEN action_name = 'DELETE' THEN 1 ELSE 0 END) AS deleted_rows
        FROM #bench_merge_actions
        """
    )
    row = cur.fetchone()
    inserted_rows = int(row[0] or 0)
    updated_rows = int(row[1] or 0)
    deleted_rows = int(row[2] or 0)
    return inserted_rows, updated_rows, deleted_rows


def build_summary_payload(
    *,
    success: bool,
    status: str,
    log_path: Path,
    source_mode: str | None = None,
    source_file: str | None = None,
    source_path: str | None = None,
    source_mtime: datetime | None = None,
    graph_item_id: str | None = None,
    graph_etag: str | None = None,
    graph_last_modified: datetime | None = None,
    sheet_name: str | None = None,
    requested_sheet: str | None = None,
    processed_sheets: list[str] | None = None,
    workbook_sheets: list[str] | None = None,
    stats: dict[str, Any] | None = None,
    inserted_rows: int = 0,
    updated_rows: int = 0,
    deleted_rows: int = 0,
    skipped_unchanged: bool = False,
    periodos: list[str] | None = None,
    columns_detected: list[str] | None = None,
    table: str = TABLE,
    error: str | None = None,
) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "success": success,
        "status": status,
        "table": table,
        "source_mode": source_mode,
        "source_file": source_file,
        "source_path": source_path,
        "source_mtime": source_mtime.isoformat() if source_mtime else None,
        "graph_item_id": graph_item_id,
        "graph_etag": graph_etag,
        "graph_last_modified": graph_last_modified.isoformat() if graph_last_modified else None,
        "sheet_name": sheet_name,
        "requested_sheet": requested_sheet,
        "processed_sheets": processed_sheets or [],
        "workbook_sheets": workbook_sheets or [],
        "rows_read": 0,
        "rows_valid": 0,
        "skipped_rows": 0,
        "inserted_rows": inserted_rows,
        "updated_rows": updated_rows,
        "deleted_rows": deleted_rows,
        "skipped_unchanged": skipped_unchanged,
        "periodos": periodos or [],
        "columns_detected": columns_detected or [],
        "log_path": str(log_path),
    }
    if stats:
        payload["rows_read"] = int(stats.get("rows_read", 0))
        payload["rows_valid"] = int(stats.get("rows_valid", 0))
        payload["skipped_rows"] = int(stats.get("rows_discarded", 0))
        payload["invalid_fecha"] = int(stats.get("invalid_fecha", 0))
        payload["invalid_cumplimiento"] = int(stats.get("invalid_cumplimiento", 0))
    if error:
        payload["error"] = error
    return payload


def run(file_path: str | None = None, periodo_override: str | None = None, sheet_override: str | None = None) -> int:
    load_env_files()
    log_path = setup_logging()
    downloaded_path: Path | None = None
    source_mtime: datetime | None = None
    source_file: str | None = None
    graph_item_id: str | None = None
    graph_etag: str | None = None
    graph_last_modified: datetime | None = None
    graph_share_url: str | None = None
    requested_sheet = sheet_override or os.getenv("BENCH_SHEET_NAME")
    workbook_sheets: list[str] = []
    processed_sheets: list[str] = []
    columns_detected: list[str] = ["fecha", "anio", "mes", "dia_habil", "negocio", "segmento", "empresa", "cumplimiento", "fecha_actualizacion"]

    try:
        graph_metadata = _graph_bench_metadata()
        graph_share_url = str(graph_metadata.get("share_url") or "").strip() or None
        graph_item_id = str(graph_metadata.get("item_id") or "").strip() or None
        graph_etag = str(graph_metadata.get("etag") or "").strip() or None
        graph_last_modified = graph_metadata.get("last_modified")
        source_file = str(graph_metadata.get("file_name") or BENCH_FILENAME).strip() or BENCH_FILENAME
        source_mtime = graph_last_modified

        if (
            graph_item_id
            and graph_etag
            and load_already_processed(graph_item_id, graph_etag)
            and bench_table_has_rows()
        ):
            logging.info(
                "Sin cambios Graph: archivo %s con item_id=%s y eTag=%s ya fue cargado.",
                source_file,
                graph_item_id,
                graph_etag,
            )
            payload = build_summary_payload(
                success=True,
                status="skipped",
                log_path=log_path,
                source_mode="graph",
                source_file=source_file,
                source_path=graph_share_url,
                source_mtime=source_mtime,
                graph_item_id=graph_item_id,
                graph_etag=graph_etag,
                graph_last_modified=graph_last_modified,
                sheet_name=None,
                requested_sheet=requested_sheet,
                processed_sheets=[],
                workbook_sheets=[],
                skipped_unchanged=True,
                periodos=[],
                columns_detected=columns_detected,
            )
            emit_payload(payload)
            return 0

        downloaded_path, graph_metadata = _download_graph_bench_file(metadata=graph_metadata)
        refresh_workbook_cache(downloaded_path)
        extracted_rows, processed_sheets, workbook_sheets = read_excel_source(downloaded_path, requested_sheet=requested_sheet)
        sheet_name = processed_sheets[0] if len(processed_sheets) == 1 else "MULTI_SHEET"
        df_prepared, stats = transform_dataframe(extracted_rows)

        if periodo_override:
            df_prepared = df_prepared[df_prepared["fecha"].astype(str).str.slice(0, 7) == str(periodo_override).strip()]
            if df_prepared.empty:
                raise RuntimeError(f"No hay filas BENCH para el periodo solicitado {periodo_override}")
            stats["rows_valid"] = len(df_prepared.index)

        with connect() as cn:
            cn.autocommit = False
            cur = cn.cursor()
            ensure_table(cur)
            inserted, updated, deleted = merge_rows(cur, df_prepared)
            cn.commit()

        write_state_file(
            {
                "schema_version": SCHEMA_VERSION,
                "source_file": source_file,
                "source_mode": "graph",
                "source_path": graph_share_url,
                "source_mtime": source_mtime.isoformat() if source_mtime else None,
                "graph_item_id": graph_item_id,
                "graph_etag": graph_etag,
                "graph_last_modified": graph_last_modified.isoformat() if graph_last_modified else None,
                "processed_at": datetime.now().replace(microsecond=0).isoformat(),
            }
        )

        payload = build_summary_payload(
            success=True,
            status="ok",
            log_path=log_path,
            source_mode="graph",
            source_file=source_file,
            source_path=graph_share_url,
            source_mtime=source_mtime,
            graph_item_id=graph_item_id,
            graph_etag=graph_etag,
            graph_last_modified=graph_last_modified,
            sheet_name=sheet_name,
            requested_sheet=requested_sheet,
            processed_sheets=processed_sheets,
            workbook_sheets=workbook_sheets,
            stats=stats,
            inserted_rows=inserted,
            updated_rows=updated,
            deleted_rows=deleted,
            periodos=safe_period_values(df_prepared),
            columns_detected=columns_detected,
        )
        logging.info("Carga BENCH completada: %s", json.dumps(payload, ensure_ascii=False, default=str))
        emit_payload(payload)
        return 0
    except Exception as exc:
        logging.exception("Fallo ETL BENCH: %s", exc)
        payload = build_summary_payload(
            success=False,
            status="error",
            log_path=log_path,
            source_mode="graph",
            source_file=source_file,
            source_path=graph_share_url,
            source_mtime=source_mtime,
            graph_item_id=graph_item_id,
            graph_etag=graph_etag,
            graph_last_modified=graph_last_modified,
            requested_sheet=requested_sheet,
            processed_sheets=processed_sheets,
            workbook_sheets=workbook_sheets,
            columns_detected=columns_detected,
            error=str(exc),
        )
        emit_payload(payload)
        return 1
    finally:
        if downloaded_path is not None:
            with suppress(Exception):
                downloaded_path.unlink(missing_ok=True)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Carga ETL BENCH Control")
    parser.add_argument("--file", required=False, help="Ruta del Excel BENCH")
    parser.add_argument("--periodo", required=False, help="Periodo YYYY-MM para filtrar la carga")
    parser.add_argument("--sheet", required=False, help="Hoja del Excel BENCH")
    args = parser.parse_args()
    sys.exit(run(file_path=args.file, periodo_override=args.periodo, sheet_override=args.sheet))
